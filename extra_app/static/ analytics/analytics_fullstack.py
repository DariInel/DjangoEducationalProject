import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np


class Vacancy:
    currency_to_rur = {
        "USD": 60.66,
        "KZT": 0.13,
        "AZN": 35.68,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "UAH": 1.64,
        "UZS": 0.0055,
        "BYR": 23.91,
        "RUR": 1,
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.get_salary_average_rur(self.salary_currency, self.salary_from, self.salary_to)
        self.area_name = vacancy['area_name']
        self.published_at = int(vacancy['published_at'][:4])

    @staticmethod
    def get_salary_average_rur(salary_currency, salary_from, salary_to):
        return (salary_from + salary_to) / 2 * Vacancy.currency_to_rur[salary_currency]


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        salary = {}
        vacancy_name_salary = {}
        t1 = {}
        city_salary = {}
        count_salary = {}
        headline = []
        number = 0
        answer1 = {}
        answer2 = {}
        answer3 = {}
        answer4 = {}
        answer5 = {}
        file_csv = open(self.file_name, encoding='utf_8_sig')
        reader = csv.reader(file_csv)
        for i, row in enumerate(reader):
            if '' not in row and len(row) == len(headline):
                current_vacancy = Vacancy(dict(zip(headline, row)))

                answer2[current_vacancy.published_at] = 1 if current_vacancy.published_at not in answer2 else answer2[
                                                                                                                  current_vacancy.published_at] + 1

                if current_vacancy.published_at in salary:
                    salary[current_vacancy.published_at].append(current_vacancy.salary_average)
                else:
                    salary[current_vacancy.published_at] = [current_vacancy.salary_average]

                for i in self.vacancy_name.split(', '):
                    if current_vacancy.name.find(i) != -1:
                        if current_vacancy.published_at in vacancy_name_salary:
                            vacancy_name_salary[current_vacancy.published_at].append(current_vacancy.salary_average)
                        else:
                            vacancy_name_salary[current_vacancy.published_at] = [current_vacancy.salary_average]

                        answer4[current_vacancy.published_at] = 1 if current_vacancy.published_at not in answer4 else \
                        answer4[current_vacancy.published_at] + 1
                        break

                if current_vacancy.area_name in city_salary:
                    city_salary[current_vacancy.area_name].append(current_vacancy.salary_average)
                else:
                    city_salary[current_vacancy.area_name] = [current_vacancy.salary_average]

                count_salary[current_vacancy.area_name] = 1 if current_vacancy.area_name not in count_salary else 1 + \
                                                                                                                  count_salary[
                                                                                                                      current_vacancy.area_name]

                number += 1

            elif i == 0:
                headline = row

        if not vacancy_name_salary:
            vacancy_name_salary = dict([(key, []) for key, value in salary.copy().items()])
            answer4 = dict([(key, 0) for key, value in answer2.copy().items()])

        for year, list_of_salaries in city_salary.items():
            answer5[year] = self.get_salary(list_of_salaries)

        for year, list_of_salaries in vacancy_name_salary.items():
            if len(list_of_salaries) != 0:
                answer3[year] = self.get_salary(list_of_salaries)
            else:
                answer3[year] = 0

        for year, list_of_salaries in salary.items():
            answer1[year] = self.get_salary(list_of_salaries)

        for year, list_of_salaries in count_salary.items():
            t1[year] = self.get_rounded(list_of_salaries, number)

        t1 = self.get_sorting(self.get_part(t1))
        answer6 = dict(t1.copy()[:10])
        t1 = dict(t1)

        answer5 = self.get_sorting(self.get_part1(t1, answer5))
        answer5 = dict(answer5[:10])

        print('Динамика уровня зарплат по годам: ' + str(answer1))
        print('Динамика количества вакансий по годам: ' + str(answer2))
        print('Динамика уровня зарплат по годам для выбранной профессии: ' + str(answer3))
        print('Динамика количества вакансий по годам для выбранной профессии: ' + str(answer4))
        print('Уровень зарплат по городам (в порядке убывания): ' + str(answer5))
        print('Доля вакансий по городам (в порядке убывания): ' + str(answer6))

        return answer1, answer2, answer3, answer4, answer5, answer6

    def get_salary(self, list_current):
        return int(sum(list_current) / len(list_current))

    def get_rounded(self, list_current, count):
        return round(list_current / count, 4)

    def get_part(self, t):
        return list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in t.items()]))

    def get_sorting(self, list_current):
        list_current.sort(key=lambda a: a[-1], reverse=True)
        return list_current

    def get_part1(self, d, list_current):
        return list(filter(lambda a: a[0] in list(d.keys()), [(key, value) for key, value in list_current.items()]))


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stats1, stats2, stats3, stats4, stats5, stats6 = dataset.csv_reader()

        report = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)
        report.generate_excel()
        report.save('report_fullstack.xlsx')
        report.generate_image()


class Report:
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def generate_excel(self):
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = font_bold
            ws2[col + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def generate_image(self):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        bar1 = ax1.bar(np.array(list(self.stats1.keys())) - 0.4, self.stats1.values(), width=0.4)
        bar2 = ax1.bar(np.array(list(self.stats1.keys())), self.stats3.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('Средняя зарплата', 'Зарплата ' + self.vacancy_name.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list(self.stats1.keys())) - 0.2, list(self.stats1.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(np.array(list(self.stats2.keys())) - 0.4, self.stats2.values(), width=0.4)
        bar2 = ax2.bar(np.array(list(self.stats2.keys())), self.stats4.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()), prop={'size': 8})
        ax2.set_xticks(np.array(list(self.stats2.keys())) - 0.2, list(self.stats2.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats5.keys()))]), list(reversed(list(self.stats5.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.stats6.values()])
        ax4.pie(list(self.stats6.values()) + [other], labels=list(self.stats6.keys()) + ['Другие'], textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph_fullstack.png')

    def save(self, filename):
        self.wb.save(filename=filename)


if __name__ == '__main__':
    InputConnect()